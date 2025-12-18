# -*- coding: utf-8 -*-
"""
ui/interfaz.py
Interfaz gráfica principal de QualityVision
Layout y componentes visuales
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import threading

from utils.constantes import COLORES, FUENTES, VENTANA, FORMATOS_IMAGEN, CONFIG_ANALISIS
from ui.estilos import aplicar_estilos_tema_industrial, crear_boton_con_icono
from ui.animaciones import GestorAnimaciones
from core.funciones import AnalizadorDefectos
from core.graficas import GeneradorGraficas
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


class QualityVisionUI:
    """Interfaz principal de la aplicación"""
    
    def __init__(self, root, usar_modelo_real=False, ruta_modelo=None):
        self.root = root
        self._configurar_ventana()
        
        # Componentes del sistema
        self.analizador = AnalizadorDefectos(
            usar_modelo_real=usar_modelo_real,
            ruta_modelo=ruta_modelo
        )
        self.generador_graficas = GeneradorGraficas()
        self.animaciones = GestorAnimaciones()
        
        # Estado de la aplicación
        self.imagen_tk = None
        self.canvas_img_id = None
        self.analisis_en_curso = False
        self.resultados_actuales = None
        self.imagen_cargada = False  # Nueva variable para saber si hay imagen
        
        # Nuevas funcionalidades
        self.historial_inspecciones = []  # Lista de dicts con resultados
        self.contadores = {
            'total_inspecciones': 0,
            'defectos_detectados': 0,
            'piezas_ok': 0,
            'tiempo_promedio': 0.0,
            'precision_promedio': 0.0
        }
        
        # Construir interfaz
        aplicar_estilos_tema_industrial()
        self._construir_interfaz()
        
        # Inicializar componentes
        self._actualizar_barra_riesgo(None, 0.0)  # Estado inicial sin análisis
        self._actualizar_contadores()
        self._actualizar_estadisticas_laterales()
        
        # Configurar drag & drop
        self._configurar_drag_drop()
        
    def _configurar_ventana(self):
        """Configura la ventana principal"""
        self.root.title("QualityVision - Sistema de Detección de Defectos Industriales")
        self.root.geometry(f"{VENTANA['ancho_inicial']}x{VENTANA['alto_inicial']}")
        self.root.minsize(VENTANA['ancho_minimo'], VENTANA['alto_minimo'])
        self.root.configure(bg=COLORES['bg_principal'])
        
        # Centrar ventana
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (VENTANA['ancho_inicial'] // 2)
        y = (self.root.winfo_screenheight() // 2) - (VENTANA['alto_inicial'] // 2)
        self.root.geometry(f"+{x}+{y}")
        
    def _construir_interfaz(self):
        """Construye todos los elementos de la interfaz"""
        # Header con título y menú
        self._crear_header()
        
        # Cuerpo principal: panel izquierdo (imagen) y derecho (resultados)
        self._crear_cuerpo()
        
        # Footer con información
        self._crear_footer()
        
    def _crear_header(self):
        """Crea el encabezado con título y menú de opciones"""
        header = tk.Frame(self.root, bg=COLORES['bg_principal'], height=70)
        header.pack(fill='x', padx=0, pady=0)
        header.pack_propagate(False)
        
        # Título con icono
        titulo_frame = tk.Frame(header, bg=COLORES['bg_principal'])
        titulo_frame.pack(side='left', padx=20, pady=15)
        
        icono_lbl = tk.Label(
            titulo_frame,
            text="🔍",
            font=('Segoe UI', 24),
            bg=COLORES['bg_principal'],
            fg=COLORES['acento_naranja']
        )
        icono_lbl.pack(side='left', padx=(0, 10))
        
        titulo_lbl = tk.Label(
            titulo_frame,
            text="QualityVision",
            font=FUENTES['titulo'],
            bg=COLORES['bg_principal'],
            fg=COLORES['texto_principal']
        )
        titulo_lbl.pack(side='left')
        
        subtitulo_lbl = tk.Label(
            titulo_frame,
            text="Sistema de Detección de Defectos Industriales",
            font=FUENTES['pequena'],
            bg=COLORES['bg_principal'],
            fg=COLORES['texto_secundario']
        )
        subtitulo_lbl.pack(side='left', padx=(15, 0))
        
        # Botón de menú (tres puntos)
        btn_menu = tk.Button(
            header,
            text="⋮",
            font=('Segoe UI', 20, 'bold'),
            bg=COLORES['bg_secundario'],
            fg=COLORES['texto_principal'],
            activebackground=COLORES['bg_hover'],
            activeforeground=COLORES['acento_naranja'],
            relief='flat',
            bd=0,
            padx=15,
            pady=5,
            cursor='hand2',
            command=self._mostrar_menu_opciones
        )
        btn_menu.pack(side='right', padx=20, pady=15)
        
        # Línea separadora
        separador = tk.Frame(self.root, bg=COLORES['borde_normal'], height=1)
        separador.pack(fill='x')
        
    def _crear_cuerpo(self):
        """Crea el cuerpo principal con paneles izquierdo, central y lateral"""
        cuerpo = tk.Frame(self.root, bg=COLORES['bg_principal'])
        cuerpo.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Panel izquierdo: Imagen (40%)
        self.panel_izquierdo = self._crear_panel_imagen(cuerpo)
        self.panel_izquierdo.pack(side='left', fill='both', expand=True, padx=(0, 8))
        
        # Panel central: Resultados (50%)
        self.panel_central = self._crear_panel_resultados(cuerpo)
        self.panel_central.pack(side='left', fill='both', expand=True, padx=(0, 8))
        
        # Panel lateral derecho: Estadísticas rápidas (10%)
        self.panel_lateral = self._crear_panel_estadisticas(cuerpo)
        self.panel_lateral.pack(side='right', fill='both', expand=False, padx=(8, 0), ipadx=10)
        
    def _crear_panel_imagen(self, parent):
        """Crea el panel de visualización de imagen"""
        panel = tk.Frame(parent, bg=COLORES['bg_secundario'], relief='flat')
        
        # Frame del título
        titulo_frame = tk.Frame(panel, bg=COLORES['bg_secundario'], height=40)
        titulo_frame.pack(fill='x', padx=15, pady=(15, 5))
        titulo_frame.pack_propagate(False)
        
        tk.Label(
            titulo_frame,
            text="📷 Imagen de Análisis",
            font=FUENTES['subtitulo'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['acento_naranja']
        ).pack(side='left', anchor='w')
        
        # Canvas para imagen con área de drop
        canvas_container = tk.Frame(panel, bg=COLORES['bg_panel'])
        canvas_container.pack(fill='both', expand=True, padx=15, pady=(10, 15))
        
        self.canvas_imagen = tk.Canvas(
            canvas_container,
            bg=COLORES['bg_panel'],
            highlightthickness=2,
            highlightbackground=COLORES['borde_normal'],
            cursor='hand2'
        )
        self.canvas_imagen.pack(fill='both', expand=True)
        
        # Texto placeholder - centrar dinámicamente
        self._centrar_placeholder()
        
        # Bind click para cargar
        self.canvas_imagen.bind('<Button-1>', lambda e: self.cargar_imagen())
        
        # Bind para redimensionar placeholder
        self.canvas_imagen.bind('<Configure>', lambda e: self._centrar_placeholder())
        
        # Controles
        self._crear_controles_imagen(panel)
        
        return panel
        
    def _centrar_placeholder(self):
        """Centra dinámicamente el texto placeholder en el canvas"""
        # No mostrar placeholder si hay imagen cargada
        if self.imagen_cargada:
            return
            
        # Eliminar placeholder anterior si existe
        if hasattr(self, 'canvas_placeholder'):
            try:
                self.canvas_imagen.delete(self.canvas_placeholder)
            except:
                pass
        
        # Obtener dimensiones del canvas
        canvas_width = self.canvas_imagen.winfo_width()
        canvas_height = self.canvas_imagen.winfo_height()
        
        # Si el canvas aún no tiene tamaño, usar valores por defecto
        if canvas_width <= 1:
            canvas_width = 700
        if canvas_height <= 1:
            canvas_height = 500
        
        # Crear texto centrado
        self.canvas_placeholder = self.canvas_imagen.create_text(
            canvas_width // 2,
            canvas_height // 2,
            text="📁 Arrastra una imagen aquí\no haz clic para seleccionar",
            font=('Segoe UI', 16, 'bold'),  # Fuente más grande
            fill=COLORES['texto_secundario'],
            justify='center',
            anchor='center'
        )
        
    def _crear_controles_imagen(self, parent):
        """Crea los controles del panel de imagen"""
        controles = tk.Frame(parent, bg=COLORES['bg_secundario'])
        controles.pack(fill='x', padx=15, pady=(0, 15))
        
        # Botones principales
        btn_frame = tk.Frame(controles, bg=COLORES['bg_secundario'])
        btn_frame.pack(fill='x', pady=(0, 10))
        
        self.btn_cargar = crear_boton_con_icono(
            btn_frame,
            "Cargar Imagen",
            self.cargar_imagen,
            'Industrial.TButton',
            '📁'
        )
        self.btn_cargar.pack(side='left', padx=(0, 8))
        
        self.btn_analizar = crear_boton_con_icono(
            btn_frame,
            "Analizar Imagen",
            self.iniciar_analisis,
            'Primary.TButton',
            '🔍'
        )
        self.btn_analizar.pack(side='left', padx=(0, 8))
        self.btn_analizar.config(state='disabled')
        
        self.btn_limpiar = crear_boton_con_icono(
            btn_frame,
            "Limpiar",
            self.limpiar_todo,
            'Industrial.TButton',
            '🗑️'
        )
        self.btn_limpiar.pack(side='left')
        
        # Barra de progreso
        self.progress = ttk.Progressbar(
            controles,
            mode='indeterminate',
            style='Industrial.Horizontal.TProgressbar'
        )
        self.progress.pack(fill='x', pady=(0, 8))
        
        # Estado
        self.lbl_estado = tk.Label(
            controles,
            text="⚡ Estado: Listo para análisis",
            font=FUENTES['pequena'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['texto_secundario'],
            anchor='w'
        )
        self.lbl_estado.pack(fill='x')
        
    def _crear_panel_resultados(self, parent):
        """Crea el panel de visualización de resultados"""
        panel = tk.Frame(parent, bg=COLORES['bg_secundario'], relief='flat')
        
        # Título y botones de acción
        titulo_frame = tk.Frame(panel, bg=COLORES['bg_secundario'], height=40)
        titulo_frame.pack(fill='x', padx=15, pady=(15, 5))
        titulo_frame.pack_propagate(False)
        
        tk.Label(
            titulo_frame,
            text="📊 Resultados del Análisis",
            font=FUENTES['subtitulo'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['acento_naranja']
        ).pack(side='left', anchor='w')
        
        # Botones de acción
        btn_frame = tk.Frame(titulo_frame, bg=COLORES['bg_secundario'])
        btn_frame.pack(side='right', anchor='e')
        
        # Botón exportar resultados
        self.btn_exportar_resultados = tk.Button(
            btn_frame,
            text="📊 Exportar Resultados",
            command=self._exportar_resultados,
            bg=COLORES['bg_panel'],  # Color gris cuando no hay datos
            fg=COLORES['texto_secundario'],
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            padx=10,
            pady=2,
            cursor='hand2',
            state='normal'  # Siempre visible
        )
        self.btn_exportar_resultados.pack(side='right', padx=(5, 0))
        
        # Botón exportar historial
        self.btn_exportar_historial = tk.Button(
            btn_frame,
            text="📋 Exportar Historial",
            command=self._exportar_historial,
            bg=COLORES['bg_panel'],  # Color gris cuando no hay datos
            fg=COLORES['texto_secundario'],
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            padx=10,
            pady=2,
            cursor='hand2',
            state='normal'  # Siempre visible
        )
        self.btn_exportar_historial.pack(side='right', padx=(5, 0))
        
        # Barra visual de riesgo
        self._crear_barra_riesgo(panel)
        
        # Notebook para tabs
        style = ttk.Style()
        style.configure('Industrial.TNotebook', background=COLORES['bg_secundario'])
        style.configure('Industrial.TNotebook.Tab', 
                       background=COLORES['bg_panel'],
                       foreground=COLORES['texto_principal'],
                       padding=[15, 8])
        style.map('Industrial.TNotebook.Tab',
                 background=[('selected', COLORES['bg_secundario'])],
                 foreground=[('selected', COLORES['acento_naranja'])])
        
        self.notebook = ttk.Notebook(panel, style='Industrial.TNotebook')
        self.notebook.pack(fill='both', expand=True, padx=15, pady=(10, 15))
        
        # Tabs
        self.tab_probabilidades = tk.Frame(self.notebook, bg=COLORES['bg_panel'])
        self.tab_mapa_calor = tk.Frame(self.notebook, bg=COLORES['bg_panel'])
        self.tab_metricas = tk.Frame(self.notebook, bg=COLORES['bg_panel'])
        self.tab_historial = tk.Frame(self.notebook, bg=COLORES['bg_panel'])
        
        self.notebook.add(self.tab_probabilidades, text='📊 Probabilidades')
        self.notebook.add(self.tab_mapa_calor, text='🔥 Mapa de Calor')
        self.notebook.add(self.tab_metricas, text='📈 Métricas')
        self.notebook.add(self.tab_historial, text='📋 Historial')
        # Frame fijo para acciones del historial (NO se borra)
        self.historial_acciones = tk.Frame(self.tab_historial, bg=COLORES['bg_panel'])
        self.historial_acciones.pack(fill='x', pady=(10, 5))

        self.btn_revision = tk.Button(
            self.tab_historial,
            text="⚠️ Ver piezas en revisión humana",
            command=self._mostrar_revision_humana,
            bg=COLORES['acento_naranja'],
            fg='white',
            font=FUENTES['boton'],
            relief='flat',
            cursor='hand2'
        )
        self.btn_revision.pack(pady=10)



        # Placeholders
        self._crear_placeholder_resultados()
        
        return panel
        
    def _crear_panel_estadisticas(self, parent):
        """Crea el panel lateral de estadísticas rápidas"""
        panel = tk.Frame(parent, bg=COLORES['bg_secundario'], relief='flat')
        
        # Título
        titulo_frame = tk.Frame(panel, bg=COLORES['bg_secundario'], height=40)
        titulo_frame.pack(fill='x', padx=10, pady=(15, 5))
        titulo_frame.pack_propagate(False)
        
        tk.Label(
            titulo_frame,
            text="📈 Estadísticas",
            font=FUENTES['subtitulo'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['acento_naranja']
        ).pack(side='left', anchor='w')
        
        # Contenedor de estadísticas
        stats_frame = tk.Frame(panel, bg=COLORES['bg_panel'])
        stats_frame.pack(fill='both', expand=True, padx=10, pady=(10, 15))
        
        # Últimas 5 inspecciones
        ultimas_frame = tk.Frame(stats_frame, bg=COLORES['bg_panel'])
        ultimas_frame.pack(fill='x', pady=(10, 5))
        
        tk.Label(
            ultimas_frame,
            text="Últimas inspecciones:",
            font=FUENTES['boton'],
            bg=COLORES['bg_panel'],
            fg=COLORES['acento_naranja']
        ).pack(anchor='w', padx=5, pady=(0, 5))
        
        self.lbl_ultimas = []
        for i in range(5):
            lbl = tk.Label(
                ultimas_frame,
                text="---",
                font=FUENTES['pequena'],
                bg=COLORES['bg_panel'],
                fg=COLORES['texto_secundario'],
                anchor='w'
            )
            lbl.pack(fill='x', padx=10, pady=1)
            self.lbl_ultimas.append(lbl)
        
        # Tasa de defectos
        tasa_frame = tk.Frame(stats_frame, bg=COLORES['bg_panel'])
        tasa_frame.pack(fill='x', pady=5)
        
        tk.Label(
            tasa_frame,
            text="Tasa de defectos:",
            font=FUENTES['boton'],
            bg=COLORES['bg_panel'],
            fg=COLORES['acento_naranja']
        ).pack(anchor='w', padx=5, pady=(0, 5))
        
        self.lbl_tasa_defectos = tk.Label(
            tasa_frame,
            text="0.0%",
            font=FUENTES['normal'],
            bg=COLORES['bg_panel'],
            fg=COLORES['acento_rojo']
        )
        self.lbl_tasa_defectos.pack(anchor='w', padx=10)
        
        # Rendimiento
        perf_frame = tk.Frame(stats_frame, bg=COLORES['bg_panel'])
        perf_frame.pack(fill='x', pady=5)
        
        tk.Label(
            perf_frame,
            text="Rendimiento:",
            font=FUENTES['boton'],
            bg=COLORES['bg_panel'],
            fg=COLORES['acento_naranja']
        ).pack(anchor='w', padx=5, pady=(0, 5))
        
        self.lbl_rendimiento = tk.Label(
            perf_frame,
            text="---",
            font=FUENTES['pequena'],
            bg=COLORES['bg_panel'],
            fg=COLORES['texto_principal']
        )
        self.lbl_rendimiento.pack(anchor='w', padx=10)
        
        # Botones de acción rápida
        acciones_frame = tk.Frame(stats_frame, bg=COLORES['bg_panel'])
        acciones_frame.pack(fill='x', pady=(10, 0))
        
        tk.Label(
            acciones_frame,
            text="Acciones rápidas:",
            font=FUENTES['boton'],
            bg=COLORES['bg_panel'],
            fg=COLORES['acento_naranja']
        ).pack(anchor='w', padx=5, pady=(0, 5))
        
        btn_batch = crear_boton_con_icono(
            acciones_frame,
            "Modo Batch",
            self._modo_batch,
            'Industrial.TButton',
            '📦'
        )
        btn_batch.pack(fill='x', padx=5, pady=2)
        
        btn_export = crear_boton_con_icono(
            acciones_frame,
            "Exportar",
            self._exportar_historial,
            'Industrial.TButton',
            '💾'
        )
        btn_export.pack(fill='x', padx=5, pady=2)
        
        return panel
        
    def _modo_batch(self):
        """Modo batch para procesar múltiples imágenes"""
        messagebox.showinfo("Modo Batch", "Funcionalidad próximamente disponible.\n\nPermitirá procesar múltiples imágenes de una carpeta.")
    
    def _exportar_historial(self):
        """Exporta el historial a CSV"""
        if not self.historial_inspecciones:
            messagebox.showwarning("Sin datos", "No hay inspecciones en el historial para exportar.")
            return
        
        from tkinter import filedialog
        import csv
        
        archivo = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Guardar historial"
        )
        
        if archivo:
            try:
                with open(archivo, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Fecha/Hora', 'Resultado', 'Confianza', 'Tiempo(s)', 'Prob_DEFECT', 'Prob_OK'])
                    
                    for entrada in self.historial_inspecciones:
                        writer.writerow([
                            entrada['fecha_hora'],
                            entrada['prediccion'],
                            f"{entrada['confianza']:.4f}",
                            f"{entrada['tiempo']:.3f}",
                            f"{entrada['probabilidades'][0]:.4f}",
                            f"{entrada['probabilidades'][1]:.4f}"
                        ])
                
                messagebox.showinfo("Exportado", f"Historial exportado a:\n{archivo}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar: {str(e)}")
        
    def _crear_barra_riesgo(self, parent):
        """Crea la barra visual de riesgo"""
        barra_frame = tk.Frame(parent, bg=COLORES['bg_secundario'])
        barra_frame.pack(fill='x', padx=15, pady=(0, 10))
        
        # Etiqueta
        tk.Label(
            barra_frame,
            text="🚨 Nivel de Riesgo:",
            font=FUENTES['normal'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['texto_principal']
        ).pack(side='left', padx=(0, 10))
        
        # Barra de progreso para riesgo
        self.barra_riesgo = ttk.Progressbar(
            barra_frame,
            orient='horizontal',
            mode='determinate',
            length=200,
            style='Industrial.Horizontal.TProgressbar'
        )
        self.barra_riesgo.pack(side='left', padx=(0, 10))
        self.barra_riesgo['value'] = 0
        
        # Etiqueta de nivel
        self.lbl_nivel_riesgo = tk.Label(
            barra_frame,
            text="BAJO",
            font=FUENTES['boton'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['acento_verde']
        )
        self.lbl_nivel_riesgo.pack(side='left')
        
    def _crear_placeholder_resultados(self):
        """Crea mensajes placeholder en los tabs"""
        for tab, mensaje in [
            (self.tab_probabilidades, "Gráfica de probabilidades por clase\naparecerá aquí tras el análisis"),
            (self.tab_mapa_calor, "Mapa de calor con áreas de interés\naparecerá aquí tras el análisis"),
            (self.tab_metricas, "Métricas detalladas del análisis\naparecerán aquí tras el análisis"),
            (self.tab_historial, "Historial de inspecciones realizadas\naparecerá aquí")
        ]:
            lbl = tk.Label(
                tab,
                text=f"⏳\n\n{mensaje}",
                font=FUENTES['normal'],
                bg=COLORES['bg_panel'],
                fg=COLORES['texto_secundario'],
                justify='center'
            )
            lbl.place(relx=0.5, rely=0.5, anchor='center')
        
            
    def _crear_footer(self):
        """Crea el footer con información y contadores en tiempo real"""
        separador = tk.Frame(self.root, bg=COLORES['borde_normal'], height=1)
        separador.pack(fill='x')
        
        footer = tk.Frame(self.root, bg=COLORES['bg_principal'], height=60)
        footer.pack(fill='x')
        footer.pack_propagate(False)
        
        # Panel izquierdo: Información del sistema
        info_frame = tk.Frame(footer, bg=COLORES['bg_principal'])
        info_frame.pack(side='left', padx=20, pady=10)
        
        tk.Label(
            info_frame,
            text="🤖 Motor: Red Neuronal Convolucional (CNN) | Versión 4.0.0",
            font=FUENTES['pequena'],
            bg=COLORES['bg_principal'],
            fg=COLORES['texto_secundario']
        ).pack(anchor='w')
        
        tk.Label(
            info_frame,
            text="© 2025 QualityVision",
            font=FUENTES['pequena'],
            bg=COLORES['bg_principal'],
            fg=COLORES['texto_secundario']
        ).pack(anchor='w')
        
        # Panel derecho: Contadores en tiempo real
        contadores_frame = tk.Frame(footer, bg=COLORES['bg_principal'])
        contadores_frame.pack(side='right', padx=20, pady=10)
        
        # Crear labels para contadores
        self.lbl_contador_total = tk.Label(
            contadores_frame,
            text="📊 Total: 0",
            font=FUENTES['pequena'],
            bg=COLORES['bg_principal'],
            fg=COLORES['acento_azul']
        )
        self.lbl_contador_total.pack(anchor='e')
        
        self.lbl_contador_defectos = tk.Label(
            contadores_frame,
            text="⚠️ Defectos: 0",
            font=FUENTES['pequena'],
            bg=COLORES['bg_principal'],
            fg=COLORES['acento_rojo']
        )
        self.lbl_contador_defectos.pack(anchor='e')
        
        self.lbl_contador_ok = tk.Label(
            contadores_frame,
            text="✅ OK: 0",
            font=FUENTES['pequena'],
            bg=COLORES['bg_principal'],
            fg=COLORES['acento_verde']
        )
        self.lbl_contador_ok.pack(anchor='e')
        
        self.lbl_tiempo_promedio = tk.Label(
            contadores_frame,
            text="⏱️ Tiempo prom.: 0.00s",
            font=FUENTES['pequena'],
            bg=COLORES['bg_principal'],
            fg=COLORES['acento_naranja']
        )
        self.lbl_tiempo_promedio.pack(anchor='e')
        
        
    # ============= MÉTODOS AUXILIARES PARA NUEVAS FUNCIONALIDADES =============
    
    def _actualizar_barra_riesgo(self, prediccion, confianza):
        """Actualiza la barra visual de riesgo"""
        if prediccion is None:
            # Estado inicial sin análisis
            riesgo = 0
            nivel = "SIN ANALIZAR"
            color = COLORES['texto_secundario']
        elif prediccion == "DEFECT":
            # Riesgo alto si es defecto con alta confianza
            if confianza >= 0.8:
                riesgo = 90
                nivel = "CRÍTICO"
                color = COLORES['acento_rojo']
            elif confianza >= 0.6:
                riesgo = 70
                nivel = "ALTO"
                color = COLORES['acento_naranja']
            else:
                riesgo = 50
                nivel = "MEDIO"
                color = COLORES['acento_naranja']
        else:
            # Riesgo bajo si es OK
            if confianza >= 0.8:
                riesgo = 10
                nivel = "BAJO"
                color = COLORES['acento_verde']
            else:
                riesgo = 30
                nivel = "MEDIO"
                color = COLORES['acento_naranja']
        
        self.barra_riesgo['value'] = riesgo
        self.lbl_nivel_riesgo.config(text=nivel, fg=color)
    
    def _actualizar_contadores(self):
        """Actualiza los labels de contadores en el footer"""
        self.lbl_contador_total.config(text=f"📊 Total: {self.contadores['total_inspecciones']}")
        self.lbl_contador_defectos.config(text=f"⚠️ Defectos: {self.contadores['defectos_detectados']}")
        self.lbl_contador_ok.config(text=f"✅ OK: {self.contadores['piezas_ok']}")
        self.lbl_tiempo_promedio.config(text=f"⏱️ Tiempo prom.: {self.contadores['tiempo_promedio']:.2f}s")
    
    def _guardar_en_historial(self, resultados):
        """Guarda los resultados en el historial"""
        from datetime import datetime
        
        entrada_historial = {
            'fecha_hora': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'prediccion': resultados['prediccion_label'],
            'confianza': resultados['confianza'],
            'tiempo': resultados['tiempo'],
            'estado_revision': resultados.get('estado_revision', 'AUTOMATICA'),
            'probabilidades': resultados['probabilidades'].tolist() if hasattr(resultados['probabilidades'], 'tolist') else resultados['probabilidades']
        }
        
        self.historial_inspecciones.append(entrada_historial)
        
        # Mantener solo las últimas 100 entradas
        if len(self.historial_inspecciones) > 100:
            self.historial_inspecciones.pop(0)
    
    def _actualizar_historial_ui(self):
        """Actualiza la pestaña de historial"""
        # Limpiar tab
        for widget in self.tab_historial.winfo_children():
            if widget is not self.btn_revision:
                widget.destroy()


        
        if not self.historial_inspecciones:
            lbl = tk.Label(
                self.tab_historial,
                text="📋 No hay inspecciones en el historial",
                font=FUENTES['normal'],
                bg=COLORES['bg_panel'],
                fg=COLORES['texto_secundario'],
                justify='center'
            )
            lbl.pack(expand=True)
            return
        
        # Crear frame principal con scrollbar
        main_frame = tk.Frame(self.tab_historial, bg=COLORES['bg_panel'])
        main_frame.pack(fill='both', expand=True)
        
        # Canvas y scrollbar
        canvas = tk.Canvas(main_frame, bg=COLORES['bg_panel'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=COLORES['bg_panel'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Header
        header_text = "Fecha/Hora\t\tResultado\tConfianza\tTiempo"
        header = tk.Label(
            scrollable_frame,
            text=header_text,
            font=FUENTES['boton'],
            bg=COLORES['bg_panel'],
            fg=COLORES['acento_naranja'],
            anchor='w'
        )
        header.pack(fill='x', padx=10, pady=5)
        
        # Separador
        sep = tk.Frame(scrollable_frame, bg=COLORES['borde_normal'], height=1)
        sep.pack(fill='x', padx=10)
        
        # Entradas del historial (últimas 20)
        for entrada in reversed(self.historial_inspecciones[-20:]):
            color = COLORES['acento_rojo'] if entrada['prediccion'] == 'DEFECT' else COLORES['acento_verde']
            
            entry_frame = tk.Frame(scrollable_frame, bg=COLORES['bg_panel'])
            entry_frame.pack(fill='x', padx=10, pady=2)
            
            # Crear texto formateado
            texto = f"{entrada['fecha_hora']}\t{entrada['prediccion']}\t\t{entrada['confianza']:.3%}\t{entrada['tiempo']:.2f}s"
            
            lbl = tk.Label(
                entry_frame,
                text=texto,
                font=FUENTES['pequena'],
                bg=COLORES['bg_panel'],
                fg=color,
                anchor='w',
                justify='left'
            )
            lbl.pack(fill='x')
        
        # Pack canvas y scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def _exportar_revision_humana_excel(self, revisiones):
        """Exporta las piezas en revisión humana a Excel"""

        if not revisiones:
            messagebox.showwarning(
                "Sin datos",
                "No hay piezas en revisión humana para exportar."
            )
            return

        from tkinter import filedialog
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Exportar revisión humana"
        )

        if not ruta:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Revision_Humana"

            # Encabezados
            headers = ["Fecha/Hora", "Resultado IA", "Confianza", "Tiempo (s)"]
            ws.append(headers)

            header_fill = PatternFill("solid", fgColor="FFC000")
            header_font = Font(bold=True)

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            # Datos
            for r in revisiones:
                ws.append([
                    r['fecha_hora'],
                    r['prediccion'],
                    round(r['confianza'] * 100, 2),
                    round(r['tiempo'], 3)
                ])

            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            wb.save(ruta)

            messagebox.showinfo(
                "Exportación exitosa",
                f"Archivo Excel creado correctamente:\n{ruta}"
            )

        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo exportar el archivo:\n{str(e)}"
            )
    
        
    def _actualizar_estadisticas_laterales(self):
        """Actualiza las estadísticas en el panel lateral"""
        # Últimas 5 inspecciones
        ultimas = self.historial_inspecciones[-5:]
        ultimas.reverse()  # Más reciente primero
        
        for i, lbl in enumerate(self.lbl_ultimas):
            if i < len(ultimas):
                entrada = ultimas[i]
                icono = "⚠️" if entrada['prediccion'] == 'DEFECT' else "✅"
                texto = f"{icono} {entrada['prediccion']} ({entrada['confianza']:.1%})"
                color = COLORES['acento_rojo'] if entrada['prediccion'] == 'DEFECT' else COLORES['acento_verde']
                lbl.config(text=texto, fg=color)
            else:
                lbl.config(text="---", fg=COLORES['texto_secundario'])
        
        # Tasa de defectos
        if self.contadores['total_inspecciones'] > 0:
            tasa = (self.contadores['defectos_detectados'] / self.contadores['total_inspecciones']) * 100
            self.lbl_tasa_defectos.config(text=f"{tasa:.1f}%")
        else:
            self.lbl_tasa_defectos.config(text="0.0%")
        
        # Rendimiento
        if self.contadores['total_inspecciones'] > 0:
            rendimiento = "Excelente" if self.contadores['tiempo_promedio'] < 1.0 else "Bueno" if self.contadores['tiempo_promedio'] < 2.0 else "Regular"
            color = COLORES['acento_verde'] if rendimiento == "Excelente" else COLORES['acento_naranja'] if rendimiento == "Bueno" else COLORES['acento_rojo']
            self.lbl_rendimiento.config(text=rendimiento, fg=color)
        else:
            self.lbl_rendimiento.config(text="---", fg=COLORES['texto_secundario'])
    
    def _modo_batch(self):
        """Abre diálogo para procesar múltiples imágenes"""
        messagebox.showinfo("Modo Batch", "Funcionalidad próximamente disponible.\nPermitirá procesar múltiples imágenes de forma automática.")
    
    def _exportar_historial(self):
        """Exporta el historial a CSV"""
        if not self.historial_inspecciones:
            messagebox.showwarning("Sin datos", "No hay inspecciones para exportar.")
            return
        
        from tkinter import filedialog
        import csv
        
        ruta = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Exportar historial"
        )
        
        if not ruta:
            return
        
        try:
            with open(ruta, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['fecha_hora', 'prediccion', 'confianza', 'tiempo', 'prob_defect', 'prob_ok']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                writer.writeheader()
                for entrada in self.historial_inspecciones:
                    writer.writerow({
                        'fecha_hora': entrada['fecha_hora'],
                        'prediccion': entrada['prediccion'],
                        'confianza': entrada['confianza'],
                        'tiempo': entrada['tiempo'],
                        'prob_defect': entrada['probabilidades'][0] if isinstance(entrada['probabilidades'], list) else entrada['probabilidades'][0],
                        'prob_ok': entrada['probabilidades'][1] if isinstance(entrada['probabilidades'], list) else entrada['probabilidades'][1]
                    })
            
            messagebox.showinfo("Exportado", f"Historial exportado a {ruta}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
        
        
    # ============= FUNCIONALIDADES =============
    
    def cargar_imagen(self):
        """Abre diálogo para cargar imagen"""
        try:
            ruta = filedialog.askopenfilename(
                title="Seleccionar imagen",
                filetypes=FORMATOS_IMAGEN
            )

            if not ruta:
                return

            print(f"📂 Intentando cargar: {ruta}")

            # Cargar con el procesador
            img, img_array = self.analizador.obtener_procesador().cargar_imagen(ruta)

            if img is None or img_array is None:
                messagebox.showerror(
                    "Error de carga",
                    "No se pudo cargar la imagen.\nVerifica el formato y intenta nuevamente."
                )
                return

            print(f"✅ Imagen cargada: {img.size}, formato: {img.format}")

            # Mostrar en canvas
            self._mostrar_imagen_en_canvas(img)

            # Actualizar estado
            nombre_archivo = ruta.split('/')[-1].split('\\')[-1]  # Para Windows y Unix
            self.lbl_estado.config(
                text=f"✅ Imagen cargada: {nombre_archivo}",
                fg=COLORES['acento_verde']
            )
            self.btn_analizar.config(state='normal')

            # Animación de pulso en botón analizar (deshabilitada para debug)
            # self.animaciones.pulso(
            #     self.btn_analizar.winfo_children()[0] if self.btn_analizar.winfo_children() else self.btn_analizar,
            #     COLORES['acento_naranja'],
            #     COLORES['acento_naranja_hover']
            # )

            print("✅ Carga de imagen completada exitosamente")

        except Exception as e:
            print(f"❌ Error en cargar_imagen: {e}")
            import traceback
            traceback.print_exc()

            messagebox.showerror(
                "Error crítico",
                f"Error al cargar la imagen:\n{str(e)}\n\n"
                "La aplicación se cerrará por seguridad."
            )
            # En lugar de cerrar automáticamente, mostrar el error y continuar
            # self.root.quit()  # NO cerrar automáticamente
        
    def _mostrar_imagen_en_canvas(self, img_pil):
        """Muestra imagen en el canvas"""
        try:
            print(f"🎨 Iniciando _mostrar_imagen_en_canvas con imagen {img_pil.size}")

            # Obtener dimensiones del canvas
            self.canvas_imagen.update_idletasks()
            canvas_w = max(self.canvas_imagen.winfo_width(), 400)
            canvas_h = max(self.canvas_imagen.winfo_height(), 300)

            print(f"📐 Canvas dimensions: {canvas_w}x{canvas_h}")

            # Redimensionar
            print("🔄 Redimensionando imagen...")
            img_resized = self.analizador.obtener_procesador().redimensionar_para_canvas(
                img_pil, canvas_w, canvas_h
            )

            print(f"🔄 Imagen redimensionada: {img_resized.size}")

            # Convertir a PhotoImage
            print("🖼️ Convirtiendo a PhotoImage...")
            self.imagen_tk = ImageTk.PhotoImage(img_resized)

            # Limpiar y mostrar
            print("🧹 Limpiando canvas...")
            self.canvas_imagen.delete('all')

            print("📍 Creando imagen en canvas...")
            self.canvas_img_id = self.canvas_imagen.create_image(
                canvas_w // 2,
                canvas_h // 2,
                image=self.imagen_tk,
                anchor='center'
            )

            # Marco decorativo
            print("🎨 Agregando marco decorativo...")
            self.canvas_imagen.create_rectangle(
                10, 10, canvas_w - 10, canvas_h - 10,
                outline=COLORES['acento_naranja'],
                width=2
            )

            # Marcar que hay imagen cargada
            self.imagen_cargada = True

            print("✅ Imagen mostrada exitosamente en canvas")

        except Exception as e:
            print(f"❌ Error en _mostrar_imagen_en_canvas: {e}")
            import traceback
            traceback.print_exc()

            messagebox.showerror(
                "Error al mostrar imagen",
                f"No se pudo mostrar la imagen:\n{str(e)}\n\n"
                "Intenta con otra imagen."
            )
    
    def iniciar_analisis(self):
        """Inicia el proceso de análisis en thread separado"""
        if self.analisis_en_curso:
            return
            
        img_array = self.analizador.obtener_procesador().obtener_array_actual()
        if img_array is None:
            messagebox.showwarning(
                "Sin imagen",
                "Por favor carga una imagen primero."
            )
            return
            
        # Cambiar estado UI
        self.analisis_en_curso = True
        self.btn_analizar.config(state='disabled')
        self.btn_cargar.config(state='disabled')
        self.btn_limpiar.config(state='disabled')
        
        self.lbl_estado.config(
            text="🔄 Analizando imagen...",
            fg=COLORES['acento_naranja']
        )
        self.progress.start(10)
        
        # Ejecutar análisis en thread
        thread = threading.Thread(target=self._realizar_analisis, daemon=True)
        thread.start()
        
    def _realizar_analisis(self):
        """Ejecuta el análisis (en thread separado)"""
        try:
            img_array = self.analizador.obtener_procesador().obtener_array_actual()
            
            # Análisis
            resultados = self.analizador.analizar_imagen(imagen_array=img_array)
            
            # Actualizar UI en thread principal
            self.root.after(0, self._finalizar_analisis, resultados)
            
        except Exception as e:
            # Manejar errores en el thread
            error_msg = f"Error durante el análisis: {str(e)}"
            print(f"❌ {error_msg}")
            import traceback
            traceback.print_exc()
            
            # Mostrar error en UI
            self.root.after(0, lambda: self._mostrar_error_analisis(error_msg))
    
    def _mostrar_error_analisis(self, error_msg):
        """Muestra error de análisis sin cerrar la aplicación"""
        self.progress.stop()
        self.analisis_en_curso = False
        self._reactivar_botones()
        
        self.lbl_estado.config(
            text="❌ Error en el análisis",
            fg=COLORES['acento_rojo']
        )
        
        messagebox.showerror(
            "Error de Análisis",
            f"No se pudo completar el análisis:\n\n{error_msg}\n\n"
            "Verifica que el modelo esté disponible o cambia a modo simulación."
        )
        
    def _finalizar_analisis(self, resultados):
        """Finaliza el análisis y muestra resultados"""
        self.progress.stop()
        self.resultados_actuales = resultados
        
        if resultados is None:
            self.lbl_estado.config(
                text="❌ Error en el análisis",
                fg=COLORES['acento_rojo']
            )
            self.analisis_en_curso = False
            self._reactivar_botones()
            return
            
        # === NUEVAS FUNCIONALIDADES ===
        # Actualizar barra de riesgo
        self._actualizar_barra_riesgo(resultados['prediccion_label'], resultados['confianza'])
        
        # Guardar en historial
        self._guardar_en_historial(resultados)
        
        # Actualizar contadores
        self.contadores['total_inspecciones'] += 1
        if resultados['prediccion_label'] == 'DEFECT':
            self.contadores['defectos_detectados'] += 1
        else:
            self.contadores['piezas_ok'] += 1
        
        # Calcular tiempo promedio
        tiempos_anteriores = [h['tiempo'] for h in self.historial_inspecciones]
        if tiempos_anteriores:
            self.contadores['tiempo_promedio'] = sum(tiempos_anteriores) / len(tiempos_anteriores)
        
        self._actualizar_contadores()
        self._actualizar_historial_ui()
        self._actualizar_estadisticas_laterales()
        
        # Mostrar resultados
        self._mostrar_resultados(resultados)
        
        # Cambiar a la pestaña de probabilidades automáticamente
        self.notebook.select(self.tab_probabilidades)
        
        # ================= DECISIÓN ASISTIDA =================
        pred = resultados['prediccion_label']
        conf = resultados['confianza']
        if conf >= 0.85:
            estado_revision = pred
        else:
            estado_revision = "REVISION_HUMANA"

        resultados['estado_revision'] = estado_revision
       # 🔥 ACTUALIZAR LA ÚLTIMA ENTRADA DEL HISTORIAL
        if self.historial_inspecciones:
            self.historial_inspecciones[-1]['estado_revision'] = estado_revision
        # ==========================
        # Zona de decisión asistida
        # ==========================

        if conf >= 0.85:
            # Decisión automática (alta confianza)
            if pred == "DEFECT":
                icono = "❌"
                color_estado = COLORES['acento_rojo']
                mensaje = f"DEFECTO CONFIRMADO ({conf:.1%})"
            else:
                icono = "✅"
                color_estado = COLORES['acento_verde']
                mensaje = f"PIEZA APROBADA ({conf:.1%})"

        elif conf >= 0.65:
            # Revisión humana recomendada (confianza media)
            icono = "⚠️"
            color_estado = COLORES['acento_naranja']
            mensaje = f"REVISIÓN HUMANA RECOMENDADA ({conf:.1%})"

        else:
            # Baja confianza (revisión obligatoria)
            icono = "⚠️"
            color_estado = COLORES['acento_naranja']
            mensaje = f"BAJA CONFIANZA – REVISIÓN OBLIGATORIA ({conf:.1%})"

 
        self.lbl_estado.config(
            text=f"{icono} ANÁLISIS COMPLETADO - {mensaje}",
            fg=color_estado
        )
        
        # Habilitar botones de exportar (cambiar colores)
        self.btn_exportar_resultados.config(
            bg=COLORES['acento_azul'],
            fg='white'
        )
        if self.historial_inspecciones:
            self.btn_exportar_historial.config(
                bg=COLORES['acento_verde'],
                fg='white'
            )
        
        self.analisis_en_curso = False
        self._reactivar_botones()
        
    def _mostrar_resultados(self, resultados):
        """Muestra los resultados en los tabs correspondientes"""
        
        # Limpiar tabs
        for tab in [self.tab_probabilidades, self.tab_mapa_calor, self.tab_metricas]:
            for widget in tab.winfo_children():
                widget.destroy()
                
        # Tab 1: Gráfica de probabilidades
        fig_probs = self.generador_graficas.crear_grafica_barras(
            resultados['probabilidades'],
            resultados['prediccion_label'],
            resultados['confianza'],
            resultados['tiempo']
        )
        canvas_probs = FigureCanvasTkAgg(fig_probs, master=self.tab_probabilidades)
        canvas_probs.draw()
        canvas_probs.get_tk_widget().pack(fill='both', expand=True)
        
        # Tab 2: Mapa de calor
        fig_heatmap = self.generador_graficas.crear_mapa_calor(
            resultados['imagen_array'],
            resultados['prediccion_idx'],
            modelo=self.analizador.obtener_modelo()
        )
        canvas_heatmap = FigureCanvasTkAgg(fig_heatmap, master=self.tab_mapa_calor)
        canvas_heatmap.draw()
        canvas_heatmap.get_tk_widget().pack(fill='both', expand=True)
        
        # Tab 3: Métricas
        fig_metricas = self.generador_graficas.crear_metricas_panel(
            resultados['probabilidades'],
            resultados['prediccion_label'],
            resultados['confianza'],
            resultados['tiempo']
        )
        canvas_metricas = FigureCanvasTkAgg(fig_metricas, master=self.tab_metricas)
        canvas_metricas.draw()
        canvas_metricas.get_tk_widget().pack(fill='both', expand=True, padx=20, pady=20)
        

    def _reactivar_botones(self):
        """Reactiva los botones después del análisis"""
        self.btn_cargar.config(state='normal')
        self.btn_limpiar.config(state='normal')
        if self.analizador.obtener_procesador().obtener_imagen_actual():
            self.btn_analizar.config(state='normal')
            
    def limpiar_todo(self):
        """Limpia toda la interfaz"""
        # Limpiar procesador
        self.analizador.obtener_procesador().limpiar()
        self.generador_graficas.limpiar()
        
        # Limpiar canvas
        self.canvas_imagen.delete('all')
        
        # Resetear estado de imagen
        self.imagen_cargada = False
        
        # Mostrar placeholder si no hay imagen
        if not self.imagen_cargada:
            # Obtener dimensiones del canvas
            canvas_width = self.canvas_imagen.winfo_width()
            canvas_height = self.canvas_imagen.winfo_height()
            
            # Si el canvas aún no tiene tamaño, usar valores por defecto
            if canvas_width <= 1:
                canvas_width = 700
            if canvas_height <= 1:
                canvas_height = 500
            
            # Crear texto centrado
            self.canvas_placeholder = self.canvas_imagen.create_text(
                canvas_width // 2,
                canvas_height // 2,
                text="📁 Arrastra una imagen aquí\no haz clic para seleccionar",
                font=('Segoe UI', 16, 'bold'),  # Fuente más grande
                fill=COLORES['texto_secundario'],
                justify='center',
                anchor='center'
            )
        
        # Limpiar tabs
        for tab in [self.tab_probabilidades, self.tab_mapa_calor, self.tab_metricas]:
            for widget in tab.winfo_children():
                widget.destroy()
        self._crear_placeholder_resultados()
        
        # Resetear estado
        self.imagen_tk = None
        self.resultados_actuales = None
        self.imagen_cargada = False  # Resetear estado de imagen
        self.lbl_estado.config(
            text="⚡ Estado: Listo para análisis",
            fg=COLORES['texto_secundario']
        )
        self.btn_analizar.config(state='disabled')
        
        # Resetear colores de botones de exportar
        self.btn_exportar_resultados.config(
            bg=COLORES['bg_panel'],
            fg=COLORES['texto_secundario']
        )
        self.btn_exportar_historial.config(
            bg=COLORES['bg_panel'],
            fg=COLORES['texto_secundario']
        )
        
    def _configurar_drag_drop(self):
        """Configura drag & drop de archivos (simplificado)"""
        # Nota: drag & drop completo requiere librerías adicionales (tkinterdnd2)
        # Por ahora, solo configuramos el visual
        def on_enter(event):
            if not self.analisis_en_curso:
                self.canvas_imagen.config(highlightbackground=COLORES['acento_naranja'])
                
        def on_leave(event):
            self.canvas_imagen.config(highlightbackground=COLORES['borde_normal'])
            
        self.canvas_imagen.bind('<Enter>', on_enter)
        self.canvas_imagen.bind('<Leave>', on_leave)
        
    def _mostrar_menu_opciones(self):
        """Muestra menú desplegable con opciones"""
        menu = tk.Menu(self.root, tearoff=0,
                      bg=COLORES['bg_panel'],
                      fg=COLORES['texto_principal'],
                      activebackground=COLORES['acento_naranja'],
                      activeforeground='white',
                      bd=0)
        
        menu.add_command(label="⚙️ Configuración", command=self._mostrar_configuracion)
        menu.add_command(label="📊 Exportar Resultados", command=self._exportar_resultados)
        menu.add_separator()
        menu.add_command(label="ℹ️ Acerca de", command=self._mostrar_acerca_de)
        menu.add_command(label="❓ Ayuda", command=self._mostrar_ayuda)
        
        # Mostrar menu en posición del mouse
        try:
            menu.tk_popup(self.root.winfo_pointerx(), self.root.winfo_pointery())
        finally:
            menu.grab_release()
            
    def _mostrar_configuracion(self):
        """Muestra ventana de configuración"""
        config_win = tk.Toplevel(self.root)
        config_win.title("Configuración")
        config_win.geometry("400x300")
        config_win.configure(bg=COLORES['bg_secundario'])
        config_win.transient(self.root)
        
        tk.Label(
            config_win,
            text="⚙️ Configuración",
            font=FUENTES['subtitulo'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['acento_naranja']
        ).pack(pady=20)
        
        tk.Label(
            config_win,
            text="Umbral de confianza: 75%\nMostrar mapa de calor: Sí\nTiempo de análisis: Auto",
            font=FUENTES['normal'],
            bg=COLORES['bg_secundario'],
            fg=COLORES['texto_principal'],
            justify='left'
        ).pack(pady=20)
        
        ttk.Button(
            config_win,
            text="Cerrar",
            command=config_win.destroy,
            style='Industrial.TButton'
        ).pack(pady=10)
        
    def _exportar_resultados(self):
        """Exporta resultados completos: imagen, mapas de calor y métricas"""
        if not self.resultados_actuales:
            messagebox.showinfo("Sin resultados", "No hay resultados para exportar.")
            return

        from tkinter import filedialog
        import os
        from PIL import Image

        # Pedir ubicación para guardar (directorio)
        directorio_base = filedialog.askdirectory(
            title="Seleccionar carpeta para guardar resultados",
            mustexist=False
        )

        if not directorio_base:
            return

        # Crear nombre de directorio con timestamp
        timestamp = self._obtener_fecha_hora().replace(':', '-').replace(' ', '_')
        directorio_resultados = os.path.join(directorio_base, f"analisis_{timestamp}")
        os.makedirs(directorio_resultados, exist_ok=True)

        try:
            r = self.resultados_actuales

            # 1. GUARDAR IMAGEN ORIGINAL
            if hasattr(self.analizador.procesador, 'imagen_actual') and self.analizador.procesador.imagen_actual:
                imagen_original = self.analizador.procesador.imagen_actual
                ruta_imagen = os.path.join(directorio_resultados, "imagen_original.jpg")
                imagen_original.save(ruta_imagen, "JPEG", quality=95)
                print(f"✅ Imagen original guardada: {ruta_imagen}")

            # 2. GUARDAR MAPA DE CALOR
            fig_heatmap = self.generador_graficas.crear_mapa_calor(
                r['imagen_array'],
                r['prediccion_idx'],
                modelo=self.analizador.obtener_modelo()
            )
            ruta_heatmap = os.path.join(directorio_resultados, "mapa_calor.png")
            fig_heatmap.savefig(ruta_heatmap, dpi=300, bbox_inches='tight', facecolor='white')
            print(f"✅ Mapa de calor guardado: {ruta_heatmap}")

            # 3. GUARDAR GRÁFICA DE PROBABILIDADES
            fig_probs = self.generador_graficas.crear_grafica_barras(
                r['probabilidades'],
                r['prediccion_label'],
                r['confianza'],
                r['tiempo']
            )
            ruta_probs = os.path.join(directorio_resultados, "probabilidades.png")
            fig_probs.savefig(ruta_probs, dpi=300, bbox_inches='tight', facecolor='white')
            print(f"✅ Gráfica de probabilidades guardada: {ruta_probs}")

            # 4. GUARDAR MÉTRICAS
            fig_metricas = self.generador_graficas.crear_metricas_panel(
                r['probabilidades'],
                r['prediccion_label'],
                r['confianza'],
                r['tiempo']
            )
            ruta_metricas = os.path.join(directorio_resultados, "metricas.png")
            fig_metricas.savefig(ruta_metricas, dpi=300, bbox_inches='tight', facecolor='white')
            print(f"✅ Métricas guardadas: {ruta_metricas}")

            # 5. GUARDAR INFORME DE TEXTO DETALLADO
            ruta_texto = os.path.join(directorio_resultados, "resultados.txt")
            texto = f"""=== RESULTADOS COMPLETOS DE ANÁLISIS - QualityVision ===

RESULTADO PRINCIPAL:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Resultado: {r['prediccion_label']}
• Confianza: {r['confianza']:.3%}
• Tiempo de análisis: {r['tiempo']:.3f}s

PROBABILIDADES DETALLADAS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• DEFECT: {r['probabilidades'][0]:.4f} ({r['probabilidades'][0]:.2%})
• OK: {r['probabilidades'][1]:.4f} ({r['probabilidades'][1]:.2%})

INFORMACIÓN TÉCNICA:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Modelo: MobileNetV2 (Transfer Learning)
• Fecha y hora: {self._obtener_fecha_hora()}
• Archivo analizado: {getattr(self.analizador.procesador, 'ruta_actual', 'N/A')}
• Modo: {'Real' if self.analizador.usar_modelo_real else 'Simulación'}

ARCHIVOS GENERADOS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• imagen_original.jpg - Imagen analizada
• mapa_calor.png - Visualización de áreas problemáticas
• probabilidades.png - Gráfica de probabilidades
• metricas.png - Métricas de rendimiento
• resultados.txt - Este informe

INTERPRETACIÓN:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{'⚠️  DEFECTO DETECTADO: La imagen contiene anomalías que requieren atención.'
 if r['prediccion_label'] == 'DEFECT'
 else '✅ PIEZA APROBADA: No se detectaron defectos significativos.'}
"""

            with open(ruta_texto, 'w', encoding='utf-8') as f:
                f.write(texto)
            print(f"✅ Informe de texto guardado: {ruta_texto}")

            # Cerrar las figuras para liberar memoria
            fig_heatmap.close()
            fig_probs.close()
            fig_metricas.close()

            messagebox.showinfo(
                "Exportación Completa",
                f"Resultados exportados exitosamente en:\n{directorio_resultados}\n\n"
                "Archivos generados:\n"
                "• imagen_original.jpg\n"
                "• mapa_calor.png\n"
                "• probabilidades.png\n"
                "• metricas.png\n"
                "• resultados.txt"
            )

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar resultados: {str(e)}")
            print(f"❌ Error en exportación: {e}")
            import traceback
            traceback.print_exc()
    
    def _obtener_fecha_hora(self):
        """Obtiene fecha y hora actual"""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


    def _mostrar_revision_humana(self):
        """Muestra solo las piezas que requieren revisión humana"""

        revisiones = [
            h for h in self.historial_inspecciones
            if h.get('estado_revision') == 'REVISION_HUMANA'
        ]

        if not revisiones:
            messagebox.showinfo(
                "Revisión Humana",
                "No hay piezas pendientes de revisión humana."
            )
            return

        win = tk.Toplevel(self.root)
        win.title("⚠️ Piezas en Revisión Humana")
        win.geometry("750x400")
        win.configure(bg=COLORES['bg_panel'])

        tk.Label(
            win,
            text="⚠️ Piezas que requieren revisión humana",
            font=FUENTES['subtitulo'],
            bg=COLORES['bg_panel'],
            fg=COLORES['acento_naranja']
        ).pack(pady=10)

        btn_exportar = tk.Button(
            win,
            text="📤 Exportar a Excel",
            command=lambda: self._exportar_revision_humana_excel(revisiones),
            bg=COLORES['acento_verde'],
            fg='white',
            font=FUENTES['boton'],
            relief='flat',
            cursor='hand2'
        )
        btn_exportar.pack(pady=(0, 10))

        frame = tk.Frame(win, bg=COLORES['bg_panel'])
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side='right', fill='y')

        lista = tk.Listbox(
            frame,
            yscrollcommand=scrollbar.set,
            font=FUENTES['pequena']
        )
        lista.pack(fill='both', expand=True)
        scrollbar.config(command=lista.yview)

        for r in revisiones:
            texto = (
                f"{r['fecha_hora']} | "
                f"Confianza: {r['confianza']:.1%} | "
                f"Resultado IA: {r['prediccion']}"
            )
            lista.insert('end', texto)







    def _mostrar_acerca_de(self):
        """Muestra información de la aplicación"""
        messagebox.showinfo(
            "Acerca de QualityVision",
            "QualityVision v1.0.0\n\n"
            "Sistema de Detección de Defectos Industriales\n"
            "Basado en Redes Neuronales Convolucionales\n\n"
            "© 2025 - Todos los derechos reservados"
        )
        
    def _mostrar_ayuda(self):
        """Muestra ayuda"""
        messagebox.showinfo(
            "Ayuda",
            "Cómo usar QualityVision:\n\n"
            "1. Carga una imagen con el botón 'Cargar Imagen'\n"
            "2. Haz clic en 'Analizar Imagen'\n"
            "3. Revisa los resultados en las pestañas:\n"
            "   - Probabilidades: Gráfica de confianza\n"
            "   - Mapa de Calor: Áreas de interés\n"
            "   - Métricas: Detalles del análisis\n\n"
            "Usa el menú (⋮) para más opciones."
        )